# -*- coding: utf-8 -*-
"""
Created on Wed Apr 20 21:37:48 2022

@author: USER
"""


"""
請抓取台南市obike 的資料

並將抓回的資料寫入到Excel表中

工作表名稱:obike

工作表的欄位標體有:
站名、地址、總數量、可借、可停、經度、緯度
"""



import json

import requests

import openpyxl

wb = openpyxl.Workbook() #建立一個空白的活頁簿
ws = wb.active #獲得目前的工作表
ws.title='obike'

ws['A1'] ='StationName'
ws['B1'] ='Address'
ws['C1'] ='Capacity'
ws['D1'] ='AvaliableBikeCount'
ws['E1']='AvaliableSpaceCount'
ws['F1']='Longitude'
ws['G1']='Latitude'

ws['A2'] = '保安轉運站'
ws['B2'] = '保安轉運站公車侯車亭旁 (文賢路一段)'
ws['C2'] = 32
ws['D2'] = 10
ws['E2'] = 22
ws['F2'] =120.230637
ws['G2']=22.932706

wb.save('excel03.xlsx')








   
    


    
    
    
    